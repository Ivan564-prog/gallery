from django.core.management import BaseCommand
from openpyxl import load_workbook
from apps.local_hierarchy.models import Region, Metropolis, Diocese, Country



class Command(BaseCommand):

    def handle(self, *args, **options):
        wb = load_workbook('/app/import_files/hierarchy.xlsx')
        sheet = wb.active
        country, _ = Country.objects.get_or_create(title='Россия')
        for row in sheet.iter_rows(min_row=3):
            region_title = row[0].value if row[0].value and len(row[0].value) > 2 else None
            metropolis_title = row[1].value if row[1].value and len(row[1].value) > 2 else None
            diocese_title = row[2].value if row[2].value and len(row[2].value) > 2 else None
            
            metropolis_title = metropolis_title or diocese_title
            if region_title and metropolis_title and diocese_title:
                region, _ = Region.objects.get_or_create(country=country, title=region_title)
                metropolis, _ = Metropolis.objects.get_or_create(region=region, title=metropolis_title)
                Diocese.objects.get_or_create(metropolis=metropolis, title=diocese_title)